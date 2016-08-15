#!/usr/bin/env python3
# Process nagios availability reports from HTML into CSV format using
# beautiful soup. Requires python3 due to dict.items() use intead of
# collections.OrderedDict.
#
# by Stefan Midjich <swehack at gmail.com> 2016-08-15

from csv import DictWriter
from argparse import ArgumentParser, FileType
from pprint import pprint as pp

from bs4 import BeautifulSoup
from openpyxl import Workbook

parser = ArgumentParser()

parser.add_argument(
    'filename',
    type=FileType('r'),
    help='HTML report file saved from Nagios.'
)

parser.add_argument(
    '--csv-output',
    type=FileType('w'),
    help='CSV Output file.'
)

parser.add_argument(
    '--excel-output',
    help='Excel filename, must be a new file as it will be created.'
)


def get_report_title(soup):
    title = soup.find('div', attrs={'class': 'dataTitle'}).get_text(strip=True)
    return title.replace('Servicegroup \'', '').rstrip('\'')


def get_host_fieldnames(soup):
    fieldnames = []
    table = soup.find('table', attrs={'class': 'data'})
    tr = table.find('tr')
    for th in tr.find_all('th', attrs={'class': 'data'}):
        text = th.get_text(strip=True)
        fieldnames.append(text)
        if text.endswith('Undetermined'):
            break
    else:
        raise ValueError('Did not find \'Undetermined\' header')

    return fieldnames


def get_service_fieldnames(soup):
    fieldnames = []
    # Get the second table found with class=data, should be the service summary
    table = soup.find_all('table', attrs={'class': 'data'})[1]
    tr = table.find('tr')
    for th in tr.find_all('th', attrs={'class': 'data'}):
        text = th.get_text(strip=True)
        fieldnames.append(text)
        if text.endswith('Undetermined'):
            break
    else:
        raise ValueError('Did not find \'Undetermined\' header')

    return fieldnames


def get_host_data(soup, fieldnames):
    rows = []
    table = soup.find('table', attrs={'class': 'data'})
    for tr in table.find_all('tr')[1:]:
        row = {}
        i = 0
        for data in tr.find_all('td'):
            row[fieldnames[i]] = data.get_text(strip=True)
            i += 1
            if i > len(fieldnames):
                break
        rows.append(row)
    
    return rows


def get_service_data(soup, fieldnames):
    rows = []
    table = soup.find_all('table', attrs={'class': 'data'})[1]
    for tr in table.find_all('tr')[1:]:
        row = {}
        i = 0
        for data in tr.find_all('td'):
            row[fieldnames[i]] = data.get_text(strip=True)

            # Last row in nagios service summary table skips the service name
            # and has a double colspan for the average data.
            if data.attrs.get('colspan', None) == '2':
                i += 1
                row[fieldnames[i]] = ''

            i += 1
            if i > len(fieldnames):
                break

        rows.append(row)
    
    return rows


def write_csv_output(csvfile, fieldnames, rows):
    writer = DictWriter(csvfile, fieldnames=fieldnames)
    writer.writeheader()

    for row in rows:
        writer.writerow(row)


def write_excel_output2(excelfile, fieldnames, rows):
    """
    This relies on xlsxwriter module
    """

    workbook = Workbook(excelfile)
    worksheet = workbook.add_worksheet()

    # First add the header
    header_col = 0
    header_row = 0

    for header in fieldnames:
        worksheet.write(header_row, header_col, header)
        header_col += 1

    # Next write the actual data rows
    header_row = 1

    for row in rows:
        header_col = 0
        for field in fieldnames:
            data = row.get(field, '')
            worksheet.set_column(header_row, header_col, len(data))
            worksheet.write(header_row, header_col, data)
            header_col += 1
        header_row += 1

    workbook.close()


def write_excel_output(worksheet, fieldnames, rows):
    """
    Relies on openpyxl module
    """

    # First add the header
    header_col = 0
    header_row = 0

    worksheet.append(fieldnames)

    for row in rows:
        cells = []
        for field in fieldnames:
            cells.append(row.get(field, ''))
        worksheet.append(cells)


def main():
    args = parser.parse_args()
    soup = BeautifulSoup(args.filename.read(), 'html.parser')
    filename = args.filename.name

    report_title = get_report_title(soup)
    report_host_fieldnames = get_host_fieldnames(soup)
    report_host_rows = get_host_data(soup, report_host_fieldnames)

    # Next get service summary info from the HTML file
    report_service_fieldnames = get_service_fieldnames(soup)
    report_service_rows = get_service_data(soup, report_service_fieldnames)

    if args.csv_output:
        # First write the host summary output
        write_csv_output(args.csv_output, report_host_fieldnames,
                         report_host_rows)

        write_csv_output(args.csv_output, report_service_fieldnames,
                         report_service_rows)

    if args.excel_output:
        wb = Workbook()
        ws1 = wb.active
        ws1.title = 'Host summary'
        write_excel_output(ws1, report_host_fieldnames, report_host_rows)

        ws2 = wb.create_sheet('Service summary')
        write_excel_output(ws2, report_service_fieldnames, report_service_rows)

        wb.save(args.excel_output)


if __name__ == '__main__':
    main()
